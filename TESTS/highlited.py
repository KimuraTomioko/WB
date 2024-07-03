import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Загрузка данных из output_all_results.xlsx
output_file = "output_all_results.xlsx"
output_df = pd.read_excel(output_file)

# Загрузка данных из stocks.xlsx
stocks_file = "stocks.xlsx"
stocks_df = pd.read_excel(stocks_file)

# Создание нового Excel файла для результата
result_wb = load_workbook(output_file)
result_sheet = result_wb.active

# Определяем стили для выделения строк
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Сравнение артикулов и выделение строк
for index, row in output_df.iterrows():
    article = row['Артикул']

    if article in stocks_df['Артикул WB'].values:
        # Находим соответствующую строку в файле output_all_results.xlsx
        result_row = index + 2  # +2 потому что openpyxl индексирует строки с 1, а pandas с 0

        # Выделяем строку цветом в новом файле
        for col in range(1, result_sheet.max_column + 1):
            result_sheet.cell(row=result_row, column=col).fill = highlight_fill

# Сохранение результата в новый файл
result_wb.save("highlighted_results.xlsx")

