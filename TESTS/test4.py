from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import re
from openpyxl.styles import PatternFill

# Функция для очистки имени файла от недопустимых символов
def clean_filename(filename):
    return re.sub(r'[^\w\s-]', '', filename)

# Функция для вычисления рекомендуемой скидки
def calculate_recommended_discount(entry_price, lowest_found_price):
    return round((1 - lowest_found_price / entry_price) * 100)

# Настройка веб-драйвера
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=service, options=options)

try:
    # Открытие страницы
    driver.get("https://www.wildberries.ru/")

    # Чтение данных из Excel файла
    input_path = 'input.xlsx'
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active

    # Создание общей электронной таблицы
    output_wb_overall = openpyxl.Workbook()
    output_sheet_overall = output_wb_overall.active
    output_sheet_overall.append(['Артикул', 'Наименование товара', 'Цена товара', 'Цена товара со скидкой', 'Ссылка товара на Wildberries', 'Рекомендуемая скидка', 'Соответствие', 'Ссылка'])

    # Проход по каждой строке в Excel (начиная с 2-й строки, чтобы пропустить заголовок)
    for row in range(2, sheet.max_row + 1):
        vendor_article = sheet.cell(row=row, column=1).value
        brand_to_search = sheet.cell(row=row, column=2).value
        name = sheet.cell(row=row, column=3).value
        wb_article = sheet.cell(row=row, column=4).value

        if not vendor_article or not brand_to_search or not name or not wb_article:
            print(f"Skipping empty row {row}")
            continue

        brand_to_search = brand_to_search.split()[0]  # Берем первое слово бренда

        wb_link = f"https://www.wildberries.ru/catalog/{wb_article}/detail.aspx"
        sheet.cell(row=row, column=5, value=wb_link)  # Запись ссылки WB в исходный файл

        # Нахождение поля ввода по ID и ввод текста
        search_input = driver.find_element(By.ID, "searchInput")
        search_input.clear()
        search_query = f"{brand_to_search} {name}"
        search_input.send_keys(search_query)
        search_input.send_keys(Keys.RETURN)
        time.sleep(10)  # Ждем загрузки результатов

        # Сохраняем ссылку на страницу парсинга
        search_page_link = driver.current_url

        product_cards = driver.find_elements(By.CSS_SELECTOR, 'article.product-card')
        if not product_cards:
            print(f"Товары с брендом '{brand_to_search}' не найдены.")
            continue

        # Создание нового Excel файла для каждого запроса
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.append(['Название товара', 'Цена товара', 'Цена со скидкой', 'Ссылка товара на Wildberries'])

        lowest_price = None
        lowest_price_item = None

        for card in product_cards:
            try:
                brand_element = card.find_element(By.CLASS_NAME, 'product-card__brand').text.strip()
                if brand_to_search.lower() in brand_element.lower():
                    name_element = card.find_element(By.CLASS_NAME, 'product-card__name').text.split('/')[1].strip()
                    try:
                        price_element = card.find_element(By.CSS_SELECTOR, '.price del').text.strip()
                    except:
                        price_element = "0"

                    try:
                        discounted_price_element = card.find_element(By.CSS_SELECTOR, '.price .price__lower-price').text.strip()
                    except:
                        discounted_price_element = "0"

                    try:
                        product_link = card.find_element(By.CSS_SELECTOR, 'a.product-card__main').get_attribute('href')
                    except:
                        product_link = card.find_element(By.CSS_SELECTOR, 'a').get_attribute('href')

                    # Записываем данные в Excel файл
                    output_sheet.append([name_element, price_element, discounted_price_element, product_link])

                    # Находим наименьшую цену со скидкой
                    discounted_price = float(re.sub(r'[^\d.]', '', discounted_price_element))
                    if lowest_price is None or discounted_price < lowest_price:
                        lowest_price = discounted_price
                        lowest_price_item = [name_element, price_element, discounted_price_element, product_link]

            except Exception as e:
                print(f"Exception occurred while parsing card: {e}")

        # Проверка на соответствие артикулов и выделение строки цветом
        for i, row_data in enumerate(output_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(wb_article) in row_data[3]:
                for cell in output_sheet[i]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                break

        # Чистим название файла и сохраняем результаты в отдельный Excel файл
        clean_name = clean_filename(f"{vendor_article}")
        output_wb.save(f"outputs_folder\\output_{clean_name}.xlsx")

        # Проверяем наличие цены в найденных данных
        entry_discounted_price = float(re.sub(r'[^\d.]', '', sheet.cell(row=row, column=6).value or "0"))
        if entry_discounted_price > 0:
            recommended_discount = calculate_recommended_discount(entry_discounted_price, lowest_price) if lowest_price is not None else 0
            correspondence = 'Да' if lowest_price is not None and entry_discounted_price <= lowest_price else 'Нет'

            # Записываем данные в общую таблицу
            output_sheet_overall.append([vendor_article, name, entry_discounted_price, lowest_price, wb_link, recommended_discount, correspondence, search_page_link])
        else:
            print(f"No price data for vendor article {vendor_article}")
            output_sheet_overall.append([vendor_article, name, entry_discounted_price, lowest_price, wb_link, 0, 'Нет', search_page_link])

    # Сохраняем общую таблицу
    output_wb_overall.save("outputs_folder\\output_overall.xlsx")
    # Сохраняем обновленный исходный файл
    wb.save('input.xlsx')

finally:
    # Закрытие драйвера
    driver.quit()
