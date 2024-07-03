from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import re

# Функция для очистки имени файла от недопустимых символов
def clean_filename(filename):
    # Удаление всех символов, кроме букв, цифр, пробелов и дефисов
    return re.sub(r'[^\w\s-]', '', filename)

# Настройка веб-драйвера
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")  # Открывать браузер в полноэкранном режиме
driver = webdriver.Chrome(service=service, options=options)

try:
    # Открытие страницы
    driver.get("https://www.wildberries.ru/")

    # Чтение данных из Excel файла
    wb = openpyxl.load_workbook('stocks.xlsx')  # Имя вашего Excel файла
    sheet = wb.active

    # Проход по каждой строке в Excel (начиная с 2-й строки, чтобы пропустить заголовок)
    for row in range(2, sheet.max_row + 1):
        brand_to_search = sheet.cell(row=row, column=2).value  # Считываем бренд из столбца B
        name = sheet.cell(row=row, column=3).value   # Считываем название из столбца C

        # Нахождение поля ввода по ID и ввод текста
        search_input = driver.find_element(By.ID, "searchInput")
        search_input.clear()  # Очистка поля ввода, если там уже что-то было

        # Вставка данных в поисковую строку
        search_query = f"{brand_to_search} {name}"
        search_input.send_keys(search_query)

        # Опционально: нажать Enter для выполнения поиска
        search_input.send_keys(Keys.RETURN)

        # Ждем немного, чтобы увидеть результаты поиска
        time.sleep(10)  # Подождать некоторое время для загрузки результатов (можно настроить)

        # Находим все карточки товаров на странице
        product_cards = driver.find_elements(By.CSS_SELECTOR, 'article.product-card')

        # Проверяем наличие карточек товаров
        if not product_cards:
            print(f"Товары с брендом '{brand_to_search}' не найдены.")
            continue  # Пропускаем текущую итерацию цикла и переходим к следующей строке в Excel

        # Создание нового Excel файла для каждого запроса
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.append(['Бренд', 'Название', 'Цена', 'Цена со скидкой', 'Артикул'])

        # Проходим по каждой карточке товара и извлекаем данные
        for card in product_cards:
            try:
                brand_element = card.find_element(By.CLASS_NAME, 'product-card__brand').text.strip()  # Бренд товара

                # Проверяем, соответствует ли текущий бренд бренду из Excel файла
                if brand_to_search.lower() in brand_element.lower():
                    name_element = card.find_element(By.CLASS_NAME, 'product-card__name').text.split('/')[0].strip()  # Название товара без слеша
                    
                    # Извлекаем цены
                    price_element = card.find_element(By.CSS_SELECTOR, '.price del').text.strip()  # Цена без скидки
                    discounted_price_element = card.find_element(By.CSS_SELECTOR, '.price .price__lower-price').text.strip()  # Цена со скидкой
                    
                    article_id = card.get_attribute('data-nm-id')  # Артикул товара

                    # Записываем данные в Excel файл
                    output_sheet.append([brand_element, name_element, price_element, discounted_price_element, article_id])

            except Exception as e:
                print(f"Exception occurred while parsing card: {e}")

        # Чистим название файла и сохраняем результаты в отдельный Excel файл
        clean_name = clean_filename(f"{brand_to_search}_{name}")
        output_wb.save(f"output_{clean_name}.xlsx")

finally:
    # Закрытие драйвера
    driver.quit()
