from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import re

# Функция для очистки имени файла от недопустимых символов
def clean_filename(filename):
    return re.sub(r'[^\w\s-]', '', filename)

# Функция для вычисления рекомендуемой скидки
def calculate_recommended_discount(entry_price, lowest_found_price):
    return round((1 - lowest_found_price / entry_price) * 100)

# Настройка веб-драйвера
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=service, options=options)

try:
    # Открытие страницы
    driver.get("https://www.wildberries.ru/")

    # Чтение данных из Excel файла
    wb = openpyxl.load_workbook('input.xlsx')
    sheet = wb.active

    # Создание общей электронной таблицы
    output_wb_overall = openpyxl.Workbook()
    output_sheet_overall = output_wb_overall.active
    output_sheet_overall.append(['Артикул', 'Наименование товара', 'Цена товара', 'Цена товара со скидкой', 'Ссылка товара на Wildberries', 'Рекомендуемая скидка', 'Соответствие'])

    # Проход по каждой строке в Excel (начиная с 2-й строки, чтобы пропустить заголовок)
    for row in range(2, sheet.max_row + 1):
        vendor_article = sheet.cell(row=row, column=1).value
        brand_to_search = sheet.cell(row=row, column=2).value.split()[0]  # Берем первое слово бренда
        name = sheet.cell(row=row, column=3).value
        wb_article = sheet.cell(row=row, column=4).value
        wb_link = f"https://www.wildberries.ru/catalog/{wb_article}/detail.aspx"

        # Нахождение поля ввода по ID и ввод текста
        search_input = driver.find_element(By.ID, "searchInput")
        search_input.clear()
        search_query = f"{brand_to_search} {name}"
        search_input.send_keys(search_query)
        search_input.send_keys(Keys.RETURN)
        time.sleep(10)  # Ждем загрузки результатов

        product_cards = driver.find_elements(By.CSS_SELECTOR, 'article.product-card')
        if not product_cards:
            print(f"Товары с брендом '{brand_to_search}' не найдены.")
            continue

        # Создание нового Excel файла для каждого запроса
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.append(['Название товара', 'Цена товара', 'Цена со скидкой', 'Ссылка товара на Wildberries'])

        lowest_price = None

        for card in product_cards:
            try:
                brand_element = card.find_element(By.CLASS_NAME, 'product-card__brand').text.strip()
                if brand_to_search.lower() in brand_element.lower():
                    name_element = card.find_element(By.CLASS_NAME, 'product-card__name').text.split('/')[1].strip()
                    try:
                        price_element = card.find_element(By.CSS_SELECTOR, '.price del').text.strip()
                    except:
                        price_element = "0"

                    try:
                        discounted_price_element = card.find_element(By.CSS_SELECTOR, '.price .price__lower-price').text.strip()
                    except:
                        discounted_price_element = "0"

                    try:
                        product_link = card.find_element(By.CSS_SELECTOR, 'a.product-card__main').get_attribute('href')
                    except:
                        product_link = card.find_element(By.CSS_SELECTOR, 'a').get_attribute('href')

                    # Записываем данные в Excel файл
                    output_sheet.append([name_element, price_element, discounted_price_element, product_link])

                    # Находим наименьшую цену со скидкой
                    discounted_price = float(re.sub(r'[^\d.]', '', discounted_price_element))
                    if lowest_price is None or discounted_price < lowest_price:
                        lowest_price = discounted_price

            except Exception as e:
                print(f"Exception occurred while parsing card: {e}")

        # Чистим название файла и сохраняем результаты в отдельный Excel файл
        clean_name = clean_filename(f"{vendor_article}")
        output_wb.save(f"output_{clean_name}.xlsx")

        # Проверяем наличие цены в входных данных
        entry_price_cell = sheet.cell(row=row, column=5).value
        if entry_price_cell is not None:
            try:
                entry_price = float(re.sub(r'[^\d.]', '', entry_price_cell))
                recommended_discount = calculate_recommended_discount(entry_price, lowest_price) if lowest_price is not None else 0
                correspondence = 'Да' if lowest_price is not None and recommended_discount > 0 else 'Нет'

                # Записываем данные в общую таблицу
                output_sheet_overall.append([vendor_article, name, entry_price, lowest_price, wb_link, recommended_discount, correspondence])
            except ValueError:
                print(f"Invalid price data for vendor article {vendor_article}")
        else:
            print(f"No price data for vendor article {vendor_article}")

    # Сохраняем общую таблицу
    output_wb_overall.save("output_overall.xlsx")

finally:
    # Закрытие драйвера
    driver.quit()
